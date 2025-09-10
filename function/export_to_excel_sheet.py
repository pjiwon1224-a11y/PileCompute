# export_to_excel_sheet.py  —  CLEAN CONSOLIDATED

from __future__ import annotations

import os
import re
import numpy as np
import pandas as pd
from typing import Dict, List, Tuple, Optional

from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

try:
    from openpyxl.utils.dataframe import dataframe_to_rows
except Exception:
    # fallback
    def dataframe_to_rows(df, index=False, header=True):
        if header:
            cols = list(df.columns)
            if index:
                cols = [df.index.name or "index"] + cols
            yield cols
        for idx, row in df.iterrows():
            vals = row.tolist()
            if index:
                vals = [idx] + vals
            yield vals


# =========================
# 공통 유틸 & 헬퍼 (한 벌만 유지)
# =========================

def _ensure_df(obj) -> Optional[pd.DataFrame]:
    """Series/리스트/딕트를 DataFrame으로 보정. 실패 시 None."""
    if obj is None:
        return None
    if isinstance(obj, pd.Series):
        return obj.to_frame().T
    if isinstance(obj, pd.DataFrame):
        return obj
    try:
        return pd.DataFrame(obj)
    except Exception:
        return None


def _write_df_grid(ws: Worksheet, df: pd.DataFrame, start_row: int, start_col: int,
                   write_header: bool = True) -> Tuple[int, int]:
    """
    ws 지정 위치에 df를 써 넣고, (end_row, end_col) 1-based 반환.
    *테이블 추가용*으로 좌표가 필요할 때 사용.
    """
    df = _ensure_df(df)
    if df is None or df.empty:
        return start_row, start_col

    r = start_row
    c0 = start_col

    # header
    if write_header:
        for j, col in enumerate(df.columns, start=c0):
            ws.cell(r, j, str(col))
        r += 1

    nrow, ncol = df.shape
    for i in range(nrow):
        for j in range(ncol):
            v = df.iat[i, j]
            if isinstance(v, np.generic):
                v = v.item()
            ws.cell(r + i, c0 + j, v)

    end_row = r + nrow - 1
    end_col = c0 + ncol - 1
    return end_row, end_col


def _write_df_append(ws: Worksheet, df: pd.DataFrame, start_row: int, start_col: int,
                     title: Optional[str] = None) -> int:
    """
    ws에 *간단히* df를 써 넣고, 다음 사용 가능한 '행 번호'를 반환.
    제목 1줄 + 헤더 1줄 + 데이터 n줄 + 공백 1줄.
    """
    r = start_row
    c = start_col

    df = _ensure_df(df)
    if df is None or df.empty:
        if title:
            ws.cell(r, c, title); r += 1
        ws.cell(r, c, "(empty)"); r += 2
        return r

    if title:
        ws.cell(r, c, title); r += 1

    # header + body
    for row in dataframe_to_rows(df, index=False, header=True):
        for j, v in enumerate(row, start=c):
            if isinstance(v, np.generic):
                v = v.item()
            ws.cell(r, j, v)
        r += 1

    r += 1
    return r


def _safe_table_name(base: str, used: set[str]) -> str:
    name = re.sub(r'[^A-Za-z0-9_]', '_', str(base)) or "T"
    if name[0].isdigit():
        name = "_" + name
    name = name[:31]
    orig = name
    k = 1
    while name in used:
        suf = f"_{k}"
        name = (orig[:(31 - len(suf))] + suf) if len(orig) + len(suf) > 31 else orig + suf
        k += 1
    used.add(name)
    return name


def _add_table(ws: Worksheet, r1: int, c1: int, r2: int, c2: int, name: str):
    if r2 < r1 or c2 < c1:
        return
    ref = f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
    tbl = Table(displayName=name, ref=ref)
    style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)
    tbl.tableStyleInfo = style
    ws.add_table(tbl)


def _autofit(ws: Worksheet, start_row: int, start_col: int, df: pd.DataFrame):
    """컬럼 너비 자동 흉내(간단)."""
    if df is None or df.empty:
        return
    for j, col in enumerate(df.columns, start=start_col):
        max_len = len(str(col))
        for i in range(df.shape[0]):
            s = "" if df.iat[i, j - start_col] is None else str(df.iat[i, j - start_col])
            if len(s) > max_len:
                max_len = len(s)
        ws.column_dimensions[get_column_letter(j)].width = min(60, max(10, int(max_len * 1.2)))


def _safe_save_xlsx(wb: Workbook, desired_path: str) -> str:
    folder = os.path.dirname(desired_path)
    if folder and not os.path.exists(folder):
        os.makedirs(folder, exist_ok=True)
    base, ext = os.path.splitext(desired_path)
    cand = desired_path
    n = 1
    while True:
        try:
            wb.save(cand)
            return cand
        except PermissionError:
            cand = f"{base}({n}){ext}"
            n += 1


# =========================
# 요약 표 빌더
# =========================

def springs_summary_from_bundle(bundle: dict) -> dict[str, pd.DataFrame]:
    """
    bundle['groups']에서 K1..K4/Kv 값을 요약 테이블로 만든다.
    return: {'고정': df, '힌지': df or None}
    """
    def _to_df(ks_norm: dict, ks_eq: dict | None):
        if not ks_norm:
            return None
        rows = []
        for key in ['Kv', 'K1', 'K2', 'K3', 'K4']:
            rows.append({
                '항목': key,
                '상시': float(ks_norm.get(key, np.nan)),
                '지진': float((ks_eq or {}).get(key, np.nan))
            })
        return pd.DataFrame(rows, columns=['항목', '상시', '지진'])

    groups = bundle.get('groups', {}) or {}
    out: Dict[str, pd.DataFrame] = {}
    if '고정' in groups and groups['고정'].get('Ks'):
        out['고정'] = _to_df(groups['고정']['Ks'].get('normal', {}),
                             groups['고정']['Ks'].get('eq', {}))
    if '힌지' in groups and groups['힌지'].get('all_out') is not None and groups['힌지'].get('Ks'):
        out['힌지'] = _to_df(groups['힌지']['Ks'].get('normal', {}),
                             groups['힌지']['Ks'].get('eq', {}))
    return out


def summarize_Ymax(results_dict: Dict[str, pd.DataFrame], ndigits: int = 3,
                   include_zero_depth: bool = True) -> pd.DataFrame:
    """
    Calc 결과 dict -> 콤보별 최대 변위 지점 요약
    - include_zero_depth=True 이면 z=0 포함
    """
    rows = []
    for combo, df in (results_dict or {}).items():
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue
        z = pd.to_numeric(df['깊이(m)'], errors='coerce').to_numpy(float)
        y = pd.to_numeric(df['변위(mm)'], errors='coerce').to_numpy(float)
        if not include_zero_depth and len(z) > 0:
            # z=0 제외
            mask = ~(np.isclose(z, 0.0))
            z = z[mask]; y = y[mask]
        if y.size == 0:
            continue
        idx = int(np.nanargmax(np.abs(y)))
        rows.append({
            '콤보': combo,
            'z_at(m)': round(float(z[idx]), ndigits),
            'y_max(mm)': round(float(y[idx]), ndigits)
        })
    return pd.DataFrame(rows).sort_values(
        '콤보',
        key=lambda s: s.astype(str).str.extract(r'(\d+)').astype(float).fillna(1e9)[0]
    ).reset_index(drop=True)

def summarize_Ymax_Mmax(results_dict: Dict[str, pd.DataFrame],
                        ndigits: int = 3,
                        include_zero_depth: bool = True) -> pd.DataFrame:
    """
    Calc 결과 dict -> 콤보별 최대 '변위'와 최대 '모멘트'를 한 표로 요약.
    - include_zero_depth=True 이면 z=0도 후보에 포함
    columns: ['콤보','z@|y|(m)','|y|max(mm)','z@|M|(m)','|M|max(kN.m)']
    """
    rows = []
    for combo, df in (results_dict or {}).items():
        if not isinstance(df, pd.DataFrame) or df.empty:
            continue

        z = pd.to_numeric(df['깊이(m)'], errors='coerce').to_numpy(float)
        y = pd.to_numeric(df['변위(mm)'], errors='coerce').to_numpy(float)
        M = pd.to_numeric(df['모멘트(kN.m)'], errors='coerce').to_numpy(float)

        # z=0 허용/제외 옵션
        if not include_zero_depth and z.size > 0:
            mask = ~np.isclose(z, 0.0)
            z = z[mask]; y = y[mask]; M = M[mask]

        if z.size == 0:
            continue

        # 최대 절대 변위, 최대 절대 모멘트
        iy = int(np.nanargmax(np.abs(y)))
        iM = int(np.nanargmax(np.abs(M)))

        rows.append({
            '콤보':         combo,
            'z@|y|(m)':     round(float(z[iy]), ndigits),
            '|y|max(mm)':   round(float(y[iy]), ndigits),
            'z@|M|(m)':     round(float(z[iM]), ndigits),
            '|M|max(kN.m)': round(float(M[iM]), ndigits),
        })

    return (pd.DataFrame(rows)
              .sort_values('콤보', key=lambda s: s.astype(str).str.extract(r'(\d+)').astype(float).fillna(1e9)[0])
              .reset_index(drop=True))






def build_pile_properties_df(bundle) -> pd.DataFrame:
    meta = bundle.get("meta", {}) or {}
    def _get(k, default=""):
        return meta.get(k, default)

    # 오버라이드 플래그(표시에 사용)
    a_src  = "수동" if meta.get("a_override")  is not None else "자동"
    Kv_src = "수동" if meta.get("Kv_override") is not None else "자동"
    Eo_src = "수동" if meta.get("Eo_override") is not None else "자동"

    rows = [
        {"항목": "말뚝공법",       "값": _get("pile_name"), "단위": ""},
        {"항목": "길이 Lp",       "값": _get("Lp"),        "단위": "m"},
        {"항목": "외경 D2",       "값": _get("D2"),        "단위": "m"},
        {"항목": "내경 D1",       "값": _get("D1"),        "단위": "m"},
        {"항목": "단면적 Ap",      "값": _get("Ap"),        "단위": "m²"},
        {"항목": "탄성계수 E",     "값": _get("E"),         "단위": "kN/m²"},
        {"항목": "단면2차모멘트 I","값": _get("I"),         "단위": "m⁴"},
        {"항목": "공법계수 a",     "값": _get("a"),         "단위": f"- ({a_src})"},
        {"항목": "수직스프링 Kv",  "값": _get("Kv"),        "단위": f"kN/m ({Kv_src})"},
        {"항목": "Eo",            "값": _get("Eo"),        "단위": f"- ({Eo_src})"},
        {"항목": "β(상시)",        "값": _get("B"),         "단위": "1/m"},
        {"항목": "β(지진)",        "값": _get("B_e"),       "단위": "1/m"},
        {"항목": "β·L(상시)",      "값": _get("BL"),        "단위": "-"},
        {"항목": "β·L(지진)",      "값": _get("BeL"),       "단위": "-"},
        {"항목": "중립축 NE",      "값": _get("NE"),        "단위": "m"},
        {"항목": "좌표(그룹)",     "값": ", ".join(map(str, _get("group_xs", []))), "단위": "m"},
        {"항목": "본수(그룹)",     "값": ", ".join(map(str, _get("counts",  []))), "단위": "EA"},
    ]
    df = pd.DataFrame(rows, columns=["항목", "값", "단위"])
    for c in ["값"]:
        df[c] = pd.to_numeric(df[c], errors="ignore")
    return df

# =========================
# 워크북 소독 (REF/그림/차트/정의된이름/유효성/외부링크)
# =========================

def sanitize_workbook_for_excel(wb: Workbook, *,
                                clear_ref_formulas: bool = True,
                                keep_values_for_formulas: bool = False,
                                remove_images: bool = False,
                                remove_charts: bool = False) -> Workbook:
    """
    - '#REF!' 수식/정의된이름/유효성/외부링크를 정리해 손상 경고를 최소화
    - remove_images=True, remove_charts=True 이면 그림/차트 제거(필요시)
    """
    # 1) 시트 내 수식/그림/차트 정리
    for ws in wb.worksheets:
        if remove_images and hasattr(ws, "_images"):
            ws._images = []
        if remove_charts and hasattr(ws, "_charts"):
            ws._charts = []

        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.startswith("="):
                    up = v.upper()
                    if "#REF!" in up:
                        cell.value = None if not keep_values_for_formulas else cell.value
                    elif clear_ref_formulas and keep_values_for_formulas:
                        # 수식을 값으로 고정하려는 의도. data_only가 아닐 땐 실제 값이 아닐 수 있음
                        cell.value = cell.value

    # 2) 정의된 이름 정리 (버전 호환)
    try:
        # dict-like
        for name, defn in list(wb.defined_names.items()):
            try:
                attr = getattr(defn, "attr_text", "")
                if attr and "#REF!" in str(attr).upper():
                    wb.defined_names.delete(name)
            except Exception:
                try:
                    if "#REF!" in str(defn).upper():
                        wb.defined_names.delete(name)
                except Exception:
                    pass
    except Exception:
        # collection-like
        try:
            for dn in list(getattr(wb.defined_names, "definedName", [])):
                try:
                    if getattr(dn, "attr_text", "") and "#REF!" in dn.attr_text.upper():
                        wb.defined_names.definedName.remove(dn)
                except Exception:
                    pass
        except Exception:
            pass

    # 3) 데이터 유효성 정리
    for ws in wb.worksheets:
        dvs = getattr(ws, "data_validations", None)
        if not dvs:
            continue
        bad = []
        for dv in list(dvs.dataValidation):
            if dv.sqref and '#REF!' in str(dv.sqref).upper():
                bad.append(dv)
        for dv in bad:
            try:
                dvs.dataValidation.remove(dv)
            except Exception:
                try:
                    ws.data_validations = None
                except Exception:
                    pass

    # 4) 외부 링크 제거
    if hasattr(wb, "_external_links"):
        try:
            wb._external_links = []
        except Exception:
            pass

    return wb


# =========================
# bundle 구조 탐색 (all_out 추출)
# =========================

def _looks_like_all_out(obj: dict) -> bool:
    if not isinstance(obj, dict):
        return False
    probe = {"summary", "table_use", "table_fact", "table_use_print", "table_fact_print"}
    return any(k in obj for k in probe)


def _extract_cases_from_bundle(bundle: dict):
    """
    bundle의 다양한 구조를 [(tag, all_out_dict), ...]로 정규화.

    지원:
      1) {"all_out": {...}}
      2) {"groups": {"고정": {"all_out": {...}}, "힌지": {"all_out": {...}}}}
      3) {"groups": {"고정": {...}, "힌지": {...}}}  # 값이 곧 all_out
      4) {"head_fix": {"all_out": {...}}, "head_hinge": {"all_out": {...}}}
      5) {"head_fix": {...}, "head_hinge": {...}}    # 값이 곧 all_out
      6) bundle 자체가 all_out
    """
    cases = []

    # 2) groups 우선
    if isinstance(bundle, dict) and "groups" in bundle and isinstance(bundle["groups"], dict):
        for tag, item in bundle["groups"].items():
            if isinstance(item, dict):
                if "all_out" in item and _looks_like_all_out(item["all_out"]):
                    cases.append((str(tag), item["all_out"]))
                elif _looks_like_all_out(item):
                    cases.append((str(tag), item))
        if cases:
            return cases

    # 4/5) head_fix / head_hinge
    if isinstance(bundle, dict):
        for tag in ("head_fix", "head_hinge"):
            if tag in bundle and isinstance(bundle[tag], dict):
                item = bundle[tag]
                label = "고정" if tag == "head_fix" else "힌지"
                if "all_out" in item and _looks_like_all_out(item["all_out"]):
                    cases.append((label, item["all_out"]))
                elif _looks_like_all_out(item):
                    cases.append((label, item))
        if cases:
            return cases

    # 1) 단일 all_out
    if isinstance(bundle, dict) and "all_out" in bundle and _looks_like_all_out(bundle["all_out"]):
        return [("default", bundle["all_out"])]

    # 6) 번들 자체가 all_out
    if _looks_like_all_out(bundle):
        return [("default", bundle)]

    raise KeyError(
        "bundle에서 all_out 구조를 찾지 못했습니다. "
        "지원: {'all_out':...}, {'groups':{...}}, {'head_fix':...,'head_hinge':...} 등. "
        f"현재 타입/키: {type(bundle)} / {list(bundle.keys()) if isinstance(bundle, dict) else 'N/A'}"
    )


# =========================
# 내보내기 1: 새 시트(표+테이블 스타일)
# =========================

def export_bundle_to_new_sheet_as_tables(
    bundle: dict,
    template_path: str,
    save_path: str,
    sheet_name: str = "계산결과(자동-표)",
    include_calc: bool = True,
    include_fix: bool = True,
    include_hinge: bool = True,
    add_summaries: bool = True,          # 스프링정수/최대변위 요약 추가
    sanitize: bool = True,
    keep_values_for_formulas: bool = False
) -> str:
    wb = load_workbook(template_path, data_only=False)
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name)

    used_names: set[str] = set()
    row, col = 1, 1

    def put_section(title: str, df: pd.DataFrame):
        nonlocal row
        df = _ensure_df(df)
        if df is None or df.empty:
            return
        ws.cell(row, col, title); row += 1
        r2, c2 = _write_df_grid(ws, df, row, col, write_header=True)
        tname = _safe_table_name(title.replace(" ", "_"), used_names)
        _add_table(ws, row, col, r2, c2, tname)
        _autofit(ws, row, col, df)
        row = r2 + 2

    # --- 말뚝 특성치 ---
    put_section("[말뚝 특성치]", build_pile_properties_df(bundle))

    # --- 반복 수렴 이력 ---
    iter_df = (bundle.get("iterations", {}) or {}).get("df_attribution")
    if isinstance(iter_df, pd.DataFrame):
        put_section("[반복 이력: 1/β, KH]", iter_df)

    # --- 스프링 정수 요약 ---
    if add_summaries:
        ss = springs_summary_from_bundle(bundle)
        if ss.get("고정") is not None:
            put_section("[스프링 정수 요약 - 고정머리]", ss["고정"])
        if ss.get("힌지") is not None:
            put_section("[스프링 정수 요약 - 힌지머리]", ss["힌지"])

    def mat_to_df(M):
        if M is None: return None
        A = np.asarray(M, float)
        df = pd.DataFrame(A)
        df.index = [f"r{i+1}" for i in range(A.shape[0])]
        return df.reset_index().rename(columns={'index': 'row'})

    groups = bundle.get("groups", {})

    # --- K/IK 행렬 ---
    if include_fix and "고정" in groups:
        km = groups["고정"]["K_mats"]
        put_section("[고정] K(상시)",  mat_to_df(km.get("K")))
        put_section("[고정] IK(상시)", mat_to_df(km.get("IK")))
        put_section("[고정] K(지진)",  mat_to_df(km.get("Ke")))
        put_section("[고정] IK(지진)", mat_to_df(km.get("IKe")))

    if include_hinge and "힌지" in groups and groups["힌지"].get("all_out") is not None:
        km = groups["힌지"]["K_mats"]
        put_section("[힌지] K(상시)",  mat_to_df(km.get("K")))
        put_section("[힌지] IK(상시)", mat_to_df(km.get("IK")))
        put_section("[힌지] K(지진)",  mat_to_df(km.get("Ke")))
        put_section("[힌지] IK(지진)", mat_to_df(km.get("IKe")))

    # --- summary/per-pile/머리력/Calc/최대변위 요약 ---
    def write_all_out(tag: str, all_out: dict):
        if not all_out: return
        put_section(f"[{tag}] summary", all_out.get("summary"))
        put_section(f"[{tag}] 사용하중 per-pile", all_out.get("table_use"))
        put_section(f"[{tag}] 사용하중 머리력 블록", all_out.get("table_use_print"))
        put_section(f"[{tag}] 계수하중 per-pile", all_out.get("table_fact"))
        put_section(f"[{tag}] 계수하중 머리력 블록", all_out.get("table_fact_print"))

        # 최대 변위 지점 요약 (z=0 포함 허용)
        ym_use  = summarize_Ymax_Mmax(all_out.get("use_results")  or {}, include_zero_depth=True)
        ym_fact = summarize_Ymax_Mmax(all_out.get("fact_results") or {}, include_zero_depth=True)
        put_section(f"[{tag}] 최대 변위/모멘트 요약 - 사용하중", ym_use)
        put_section(f"[{tag}] 최대 변위/모멘트 요약 - 계수하중", ym_fact)

        if include_calc:
            for combo, df in (all_out.get("use_results") or {}).items():
                put_section(f"[{tag}] 사용하중 Calc – {combo}", df)
            for combo, df in (all_out.get("fact_results") or {}).items():
                put_section(f"[{tag}] 계수하중 Calc – {combo}", df)

    if include_fix and "고정" in groups:
        write_all_out("고정", groups["고정"]["all_out"])
    if include_hinge and "힌지" in groups and groups["힌지"].get("all_out") is not None:
        write_all_out("힌지", groups["힌지"]["all_out"])

    # 소독 (손상 경고 완화)
    if sanitize:
        sanitize_workbook_for_excel(
            wb,
            clear_ref_formulas=True,
            keep_values_for_formulas=keep_values_for_formulas,
            remove_images=True,     # 필요에 따라 False
            remove_charts=True      # 필요에 따라 False
        )

    return _safe_save_xlsx(wb, save_path)


# =========================
# 내보내기 2: 단일 시트(심플 append)
# =========================

def export_bundle_to_single_sheet(
    bundle: dict,
    template_path: str,
    save_path: str,
    sheet_name: str = "결과",
    include_calc: bool = True,
    include_fix: bool = True,
    include_hinge: bool = True,
    sanitize: bool = True,
    keep_values_for_formulas: bool = False
) -> str:
    wb = load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.create_sheet(sheet_name)

    r = 1

    # 말뚝 특성치
    r = _write_df_append(ws, build_pile_properties_df(bundle), r, 1, title="말뚝 특성치")

    # 반복 수렴 이력
    iter_df = (bundle.get("iterations", {}) or {}).get("df_attribution")
    if isinstance(iter_df, pd.DataFrame) and not iter_df.empty:
        r = _write_df_append(ws, iter_df, r, 1, title="[반복 수렴 이력: 1/β, KH]")

    groups = bundle.get("groups", {})

    def write_all_out_block(tag: str, all_out: dict):
        nonlocal r
        if not all_out:
            return
        r = _write_df_append(ws, all_out.get("summary"), r, 1, title=f"[{tag}] summary")
        r = _write_df_append(ws, all_out.get("table_use"), r, 1, title=f"[{tag}] 사용하중 per-pile")
        r = _write_df_append(ws, all_out.get("table_use_print"), r, 1, title=f"[{tag}] 사용하중 머리력 블록")
        r = _write_df_append(ws, all_out.get("table_fact"), r, 1, title=f"[{tag}] 계수하중 per-pile")
        r = _write_df_append(ws, all_out.get("table_fact_print"), r, 1, title=f"[{tag}] 계수하중 머리력 블록")

        # 최대 변위 요약
        y_use  = summarize_Ymax(all_out.get("use_results")  or {}, include_zero_depth=True)
        y_fact = summarize_Ymax(all_out.get("fact_results") or {}, include_zero_depth=True)
        r = _write_df_append(ws, y_use,  r, 1, title=f"[{tag}] 최대 변위 요약 - 사용하중")
        r = _write_df_append(ws, y_fact, r, 1, title=f"[{tag}] 최대 변위 요약 - 계수하중")

        if include_calc:
            ws.cell(r, 1, f"[{tag}] Calc 결과표(사용하중)"); r += 1
            for combo, df in (all_out.get("use_results") or {}).items():
                r = _write_df_append(ws, df, r, 1, title=combo)
            ws.cell(r, 1, f"[{tag}] Calc 결과표(계수하중)"); r += 1
            for combo, df in (all_out.get("fact_results") or {}).items():
                r = _write_df_append(ws, df, r, 1, title=combo)

    if include_fix and "고정" in groups:
        ws.cell(r, 1, "=== 머리 고정 ==="); r += 1
        write_all_out_block("고정", groups["고정"]["all_out"])

    if include_hinge and "힌지" in groups and groups["힌지"].get("all_out") is not None:
        ws.cell(r, 1, "=== 머리 힌지 ==="); r += 1
        write_all_out_block("힌지", groups["힌지"]["all_out"])

    if sanitize:
        sanitize_workbook_for_excel(
            wb,
            clear_ref_formulas=True,
            keep_values_for_formulas=keep_values_for_formulas,
            remove_images=True,
            remove_charts=True
        )

    return _safe_save_xlsx(wb, save_path)


# =========================
# 내보내기 3: 기존 시트의 특정 영역 아래에 붙이기
# =========================

def export_bundle_to_traits_section(
    bundle: dict,
    template_path: str,
    save_path: str,
    sheet_name: str = "결과",
    include_calc: bool = True,
    include_fix: bool = True,
    include_hinge: bool = True,
    anchor_text: str = "말뚝 특성치",
    sanitize: bool = True,
    keep_values_for_formulas: bool = False
) -> str:
    wb = load_workbook(template_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    # 앵커 찾기
    anchor_cell = None
    for row in ws.iter_rows():
        for cell in row:
            if str(cell.value).strip() == anchor_text:
                anchor_cell = cell; break
        if anchor_cell: break
    anchor_row = anchor_cell.row if anchor_cell else 1

    # 앵커 아래 정리
    maxr = ws.max_row
    if maxr > anchor_row:
        ws.delete_rows(anchor_row + 1, maxr - anchor_row)

    r = anchor_row + 1

    # 말뚝 특성치
    r = _write_df_append(ws, build_pile_properties_df(bundle), r, 1, title="[말뚝 특성치]")

    # 반복 수렴 이력
    iter_df = (bundle.get("iterations", {}) or {}).get("df_attribution")
    if isinstance(iter_df, pd.DataFrame) and not iter_df.empty:
        r = _write_df_append(ws, iter_df, r, 1, title="[반복 수렴 이력: 1/β, KH]")

    groups = bundle.get("groups", {})

    def _block(tag: str, all_out: dict):
        nonlocal r
        if not all_out: return
        r = _write_df_append(ws, all_out.get("summary"), r, 1, title=f"[{tag}] summary")
        r = _write_df_append(ws, all_out.get("table_use"), r, 1, title=f"[{tag}] 사용하중 per-pile")
        r = _write_df_append(ws, all_out.get("table_use_print"), r, 1, title=f"[{tag}] 사용하중 머리력 블록")
        r = _write_df_append(ws, all_out.get("table_fact"), r, 1, title=f"[{tag}] 계수하중 per-pile")
        r = _write_df_append(ws, all_out.get("table_fact_print"), r, 1, title=f"[{tag}] 계수하중 머리력 블록")

        # 최대 변위 요약
        y_use  = summarize_Ymax(all_out.get("use_results")  or {}, include_zero_depth=True)
        y_fact = summarize_Ymax(all_out.get("fact_results") or {}, include_zero_depth=True)
        r = _write_df_append(ws, y_use,  r, 1, title=f"[{tag}] 최대 변위 요약 - 사용하중")
        r = _write_df_append(ws, y_fact, r, 1, title=f"[{tag}] 최대 변위 요약 - 계수하중")

        if include_calc:
            ws.cell(r, 1, f"[{tag}] Calc 결과표(사용하중)"); r += 1
            for combo, df in (all_out.get("use_results") or {}).items():
                r = _write_df_append(ws, df, r, 1, title=combo)
            ws.cell(r, 1, f"[{tag}] Calc 결과표(계수하중)"); r += 1
            for combo, df in (all_out.get("fact_results") or {}).items():
                r = _write_df_append(ws, df, r, 1, title=combo)

    if include_fix and "고정" in groups:
        _block("고정", groups["고정"]["all_out"])
    if include_hinge and "힌지" in groups and groups["힌지"].get("all_out") is not None:
        _block("힌지", groups["힌지"]["all_out"])

    if sanitize:
        sanitize_workbook_for_excel(
            wb,
            clear_ref_formulas=True,
            keep_values_for_formulas=keep_values_for_formulas,
            remove_images=True,
            remove_charts=True
        )

    return _safe_save_xlsx(wb, save_path)


# =========================
# 내보내기 4: OUT_* 여러 시트로 분류 출력
# =========================

def export_bundle_to_template_auto(
    bundle: dict,
    template_path: str,
    save_path: str,
    target_sheet: str | None = None,
    also_write_calc_result: bool = True
) -> str:
    wb = load_workbook(template_path)

    # 말뚝 특성치
    if "OUT_PileProps" in wb.sheetnames:
        wb.remove(wb["OUT_PileProps"])
    ws = wb.create_sheet("OUT_PileProps")
    r = 1
    ws.cell(r, 1, "말뚝 특성치"); r += 1
    for row in dataframe_to_rows(build_pile_properties_df(bundle), index=False, header=True):
        for j, v in enumerate(row, start=1):
            ws.cell(r, j, v)
        r += 1

    # OUT_* 시트 새로 만들기
    for name in ("OUT_Summary", "OUT_PerPile_Use", "OUT_PerPile_Fact", "OUT_Calc_Use", "OUT_Calc_Fact"):
        if name in wb.sheetnames:
            wb.remove(wb[name])
    ws_sum  = wb.create_sheet("OUT_Summary")
    ws_pu   = wb.create_sheet("OUT_PerPile_Use")
    ws_pf   = wb.create_sheet("OUT_PerPile_Fact")
    ws_cu   = wb.create_sheet("OUT_Calc_Use")
    ws_cf   = wb.create_sheet("OUT_Calc_Fact")

    cases = _extract_cases_from_bundle(bundle)
    r_sum = r_pu = r_pf = r_cu = r_cf = 1

    for tag, all_out in cases:
        # summary
        if isinstance(all_out.get("summary"), pd.DataFrame):
            ws_sum.cell(r_sum, 1, f"[{tag}] summary"); r_sum += 1
            for row in dataframe_to_rows(all_out["summary"].reset_index(), index=False, header=True):
                for j, v in enumerate(row, start=1):
                    ws_sum.cell(r_sum, j, v)
                r_sum += 1
            r_sum += 1

        # per-pile
        if isinstance(all_out.get("table_use"), pd.DataFrame):
            ws_pu.cell(r_pu, 1, f"[{tag}] per-pile(사용)"); r_pu += 1
            for row in dataframe_to_rows(all_out["table_use"], index=False, header=True):
                for j, v in enumerate(row, start=1):
                    ws_pu.cell(r_pu, j, v)
                r_pu += 1
            r_pu += 1

        if isinstance(all_out.get("table_fact"), pd.DataFrame):
            ws_pf.cell(r_pf, 1, f"[{tag}] per-pile(계수)"); r_pf += 1
            for row in dataframe_to_rows(all_out["table_fact"], index=False, header=True):
                for j, v in enumerate(row, start=1):
                    ws_pf.cell(r_pf, j, v)
                r_pf += 1
            r_pf += 1

        # calc
        if also_write_calc_result:
            if isinstance(all_out.get("use_results"), dict):
                ws_cu.cell(r_cu, 1, f"[{tag}] Calc(사용)"); r_cu += 1
                for combo, df in all_out["use_results"].items():
                    if isinstance(df, pd.DataFrame) and not df.empty:
                        ws_cu.cell(r_cu, 1, str(combo)); r_cu += 1
                        for row in dataframe_to_rows(df, index=False, header=True):
                            for j, v in enumerate(row, start=1):
                                ws_cu.cell(r_cu, j, v)
                            r_cu += 1
                        r_cu += 1

            if isinstance(all_out.get("fact_results"), dict):
                ws_cf.cell(r_cf, 1, f"[{tag}] Calc(계수)"); r_cf += 1
                for combo, df in all_out["fact_results"].items():
                    if isinstance(df, pd.DataFrame) and not df.empty:
                        ws_cf.cell(r_cf, 1, str(combo)); r_cf += 1
                        for row in dataframe_to_rows(df, index=False, header=True):
                            for j, v in enumerate(row, start=1):
                                ws_cf.cell(r_cf, j, v)
                            r_cf += 1
                        r_cf += 1

    return _safe_save_xlsx(wb, save_path)
